import os
import re
import requests
from tqdm import tqdm
from openpyxl import load_workbook

link_excel = 'etf_links(SZ).xlsx'
txt_folder1 = 'etf_txts深1'
txt_folder2 = 'etf_txts深2'
os.makedirs(txt_folder1, exist_ok=True)
os.makedirs(txt_folder2, exist_ok=True)

def clean_etf_code(raw):
    match = re.search(r'(ETF)?(\d{6})', raw)
    return match.group(2) if match else raw

def download_txt_files():
    wb = load_workbook(link_excel)
    ws = wb.active
    success_count = 0
    for row in tqdm(ws.iter_rows(min_row=2, values_only=True), desc='下载TXT文件'):
        etf_code_raw, url = row
        etf_code = clean_etf_code(etf_code_raw)
        if not etf_code or 'opencode=' not in str(url):
            continue
        try:
            match = re.search(r'opencode=([A-Z0-9]+\.txt)', url)
            if not match:
                continue
            filename = match.group(1)
            txt_url = f'https://reportdocs.static.szse.cn/files/text/etf/{filename}'
            resp = requests.get(txt_url, timeout=10)
            if resp.status_code == 200:
                folder = txt_folder1 if success_count < 300 else txt_folder2
                save_path = os.path.join(folder, f'{etf_code}.txt')
                with open(save_path, 'w', encoding='utf-8', errors='ignore') as f:
                    f.write(resp.text)
                success_count += 1
        except Exception as e:
            print(f"[ERROR] 下载失败: {etf_code} - {e}")

if __name__ == '__main__':
    download_txt_files()

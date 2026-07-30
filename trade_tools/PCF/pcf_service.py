from __future__ import annotations

import asyncio
import csv
import os
import re
import sys
import threading
import time
import xml.etree.ElementTree as ElementTree
import zipfile
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Callable, Iterable, Sequence
from urllib.parse import parse_qs, unquote, urljoin, urlparse

import openpyxl
import requests
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright


SSE_PAGE_URL = "https://www.sse.com.cn/disclosure/fund/etflist/"
SZSE_PAGE_URL = "https://www.szse.cn/disclosure/fund/currency/index.html"
SZSE_TEXT_BASE_URL = "https://reportdocs.static.szse.cn/files/text/etf/"
MARKETS = ("SH", "SZ")
SSE_STOCK_ETF_CATEGORIES = {
    1: "单市场股票（沪）ETF",
    2: "跨市场股票（沪深京）ETF",
    3: "跨市场股票（沪港深京）ETF",
    4: "单市场股票（科创板）ETF",
    5: "跨市场股票（含科创板）ETF",
}

ProgressCallback = Callable[[str], None]
StructuredProgressCallback = Callable[[dict[str, object]], None]
PCF_LOG_LOCK = threading.Lock()


@dataclass(frozen=True)
class LinkRecord:
    etf_code: str
    url: str


@dataclass(frozen=True)
class WhitelistRecord:
    etf_code: str
    stock_code: str
    stock_name: str
    market: str
    substitute_flag: str
    source_file: str


@dataclass(frozen=True)
class PcfComponent:
    security_code: str
    security_name: str
    quantity: str
    substitute_flag: str
    creation_premium_rate: str
    redemption_discount_rate: str
    substitution_cash_amount: str
    creation_substitution_amount: str
    redemption_substitution_amount: str
    component_market: str


PROCESSED_COMPONENT_COLUMNS = (
    ("security_code", "证券代码"),
    ("security_name", "证券名称"),
    ("quantity", "数量"),
    ("substitute_flag", "替代标志"),
    ("creation_premium_rate", "申购保证金率"),
    ("redemption_discount_rate", "赎回保证金率"),
    ("substitution_cash_amount", "现金替代金额"),
    ("creation_substitution_amount", "申购替代金额"),
    ("redemption_substitution_amount", "赎回替代金额"),
    ("component_market", "成分市场"),
)


class RequestThrottle:
    def __init__(self, delay_seconds: float) -> None:
        self.delay_seconds = max(0.0, delay_seconds)
        self._lock = threading.Lock()
        self._next_request_at = 0.0

    def wait(self) -> None:
        with self._lock:
            remaining = self._next_request_at - time.monotonic()
            if remaining > 0:
                time.sleep(remaining)
            self._next_request_at = time.monotonic() + self.delay_seconds

    def penalize(self, seconds: float) -> None:
        with self._lock:
            self._next_request_at = max(
                self._next_request_at,
                time.monotonic() + max(0.0, seconds),
            )


def emit(message: str, callback: ProgressCallback | None = None) -> None:
    if callback:
        callback(message)
    else:
        print(message, flush=True)


def report_progress(
    callback: StructuredProgressCallback | None,
    **progress: object,
) -> None:
    if callback:
        callback(progress)


def use_system_proxy() -> bool:
    return os.environ.get("PCF_USE_SYSTEM_PROXY", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }


def browser_launch_options(headless: bool) -> dict[str, object]:
    options: dict[str, object] = {"headless": headless}
    if not use_system_proxy():
        options["args"] = ["--no-proxy-server"]
    return options


def default_work_dir() -> Path:
    configured = os.environ.get("PCF_DATA_DIR", "").strip()
    if configured:
        return Path(configured).expanduser().resolve()
    if os.name == "nt":
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
    elif sys.platform == "darwin":
        base = Path.home() / "Library" / "Application Support"
    else:
        base = Path(os.environ.get("XDG_DATA_HOME", Path.home() / ".local" / "share"))
    return (base / "PrivateFundIPOsAssistant" / "pcf").resolve()


def resolve_work_dir(value: str | os.PathLike[str] | None = None) -> Path:
    path = Path(value).expanduser().resolve() if value else default_work_dir()
    path.mkdir(parents=True, exist_ok=True)
    return path


def pcf_log_path(work_dir: str | os.PathLike[str] | None = None) -> Path:
    path = resolve_work_dir(work_dir) / "logs" / "pcf.log"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def append_pcf_log(
    work_dir: str | os.PathLike[str] | None,
    message: str,
) -> None:
    line = f"{time.strftime('%Y-%m-%d %H:%M:%S')} {message}\n"
    with PCF_LOG_LOCK:
        with pcf_log_path(work_dir).open("a", encoding="utf-8") as handle:
            handle.write(line)


def normalize_market(value: str) -> str:
    market = value.strip().upper()
    if market not in MARKETS:
        raise ValueError(f"不支持的市场: {value}")
    return market


def normalize_stock_code(value: str) -> tuple[str, str]:
    raw = str(value or "").strip().upper()
    match = re.search(r"(?<!\d)(\d{6})(?:\.(SH|SZ))?(?!\d)", raw)
    if not match:
        raise ValueError("股票代码必须包含 6 位数字，例如 688825.SH")
    code, suffix = match.group(1), match.group(2) or ""
    return code, suffix


def clean_etf_code(value: object) -> str:
    match = re.search(r"(?:ETF)?(\d{6})", str(value or ""), re.IGNORECASE)
    return match.group(1) if match else ""


def market_dir(work_dir: Path, market: str) -> Path:
    path = work_dir / normalize_market(market)
    path.mkdir(parents=True, exist_ok=True)
    return path


def links_path(work_dir: Path, market: str) -> Path:
    return market_dir(work_dir, market) / "etf_links.xlsx"


def text_dir(work_dir: Path, market: str) -> Path:
    path = market_dir(work_dir, market) / "pcf_texts"
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_links(records: Sequence[LinkRecord], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ETF链接"
    sheet.append(["ETF代码", "下载链接"])
    for record in records:
        sheet.append([record.etf_code, record.url])
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 90
    workbook.save(output_path)
    return output_path


def read_links(input_path: Path) -> list[LinkRecord]:
    if not input_path.exists():
        raise FileNotFoundError(f"链接清单不存在: {input_path}")
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    records: list[LinkRecord] = []
    seen: set[tuple[str, str]] = set()
    for row in workbook.active.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        code = clean_etf_code(row[0])
        url = str(row[1] or "").strip()
        key = (code, url)
        if code and url and key not in seen:
            seen.add(key)
            records.append(LinkRecord(code, url))
    workbook.close()
    return records


async def _scrape_szse(headless: bool, callback: ProgressCallback | None) -> list[LinkRecord]:
    records: list[LinkRecord] = []
    seen: set[tuple[str, str]] = set()
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(**browser_launch_options(headless))
        page = await browser.new_page()
        response = await page.goto(SZSE_PAGE_URL, wait_until="load", timeout=60_000)
        if response and response.status >= 400:
            await browser.close()
            raise RuntimeError(f"深交所页面返回 HTTP {response.status}")
        await page.wait_for_selector("table tbody tr", state="attached", timeout=60_000)
        page_number = 1
        while True:
            rows = await page.query_selector_all("table tbody tr")
            for row in rows:
                anchors = await row.query_selector_all("td a")
                if len(anchors) < 2:
                    continue
                raw_code = (await anchors[0].inner_text()).strip()
                href = await anchors[1].get_attribute("href")
                code = clean_etf_code(raw_code)
                if code and href:
                    full_url = urljoin(SZSE_PAGE_URL, href)
                    key = (code, full_url)
                    if key not in seen:
                        seen.add(key)
                        records.append(LinkRecord(code, full_url))
            emit(f"深交所链接第 {page_number} 页，累计 {len(records)} 条", callback)
            next_button = await page.query_selector("li.next")
            if not next_button:
                break
            class_name = (await next_button.get_attribute("class") or "").lower()
            if "disabled" in class_name:
                break
            next_link = await next_button.query_selector("a")
            if not next_link:
                break
            await next_link.click()
            await page.wait_for_timeout(1_500)
            page_number += 1
        await browser.close()
    return records


async def _scrape_sse(headless: bool, callback: ProgressCallback | None) -> list[LinkRecord]:
    records: list[LinkRecord] = []
    seen: set[tuple[str, str]] = set()
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(**browser_launch_options(headless))
        page = await browser.new_page()
        response = await page.goto(SSE_PAGE_URL, wait_until="domcontentloaded", timeout=60_000)
        if response and response.status >= 400:
            await browser.close()
            raise RuntimeError(
                f"上交所页面返回 HTTP {response.status}。该站点可能限制当前网络出口，请稍后重试。"
            )
        await page.wait_for_selector("#tab_main", timeout=60_000)
        for subtype, category_name in SSE_STOCK_ETF_CATEGORIES.items():
            previous_first_code = ""
            if subtype > 1:
                current_cell = await page.query_selector(
                    "#tab_main table tbody tr:first-child td:nth-child(1)"
                )
                if current_cell:
                    previous_first_code = clean_etf_code(await current_cell.inner_text())
            await page.click(f'//*[@id="tab_main"]/div[1]/div[1]/button[{subtype}]')
            if previous_first_code:
                await page.wait_for_function(
                    r"""previous => {
                        const cell = document.querySelector(
                            '#tab_main table tbody tr:first-child td:nth-child(1)'
                        );
                        const match = cell && cell.textContent.match(/\d{6}/);
                        return match && match[0] !== previous;
                    }""",
                    arg=previous_first_code,
                    timeout=30_000,
                )
                await page.wait_for_timeout(600)
            else:
                await page.wait_for_timeout(1_200)

            page_text = await page.locator("body").inner_text()
            total_matches = re.findall(r"共\s*(\d+)\s*条", page_text)
            expected_total = int(total_matches[-1]) if total_matches else 0
            category_codes: set[str] = set()
            page_signatures: set[tuple[str, ...]] = set()
            page_number = 1

            while True:
                await page.wait_for_selector("#tab_main table tbody tr", timeout=30_000)
                rows = await page.query_selector_all("#tab_main table tbody tr")
                page_codes: list[str] = []
                for row in rows:
                    code_element = await row.query_selector(":scope > td:nth-child(1)")
                    link_element = await row.query_selector(":scope > td:nth-child(7) a")
                    if not code_element or not link_element:
                        continue
                    code = clean_etf_code(await code_element.inner_text())
                    href = await link_element.get_attribute("href")
                    if code and href:
                        page_codes.append(code)
                        category_codes.add(code)
                        full_url = urljoin(SSE_PAGE_URL, href.replace("https:/query", "https://query"))
                        key = (code, full_url)
                        if key not in seen:
                            seen.add(key)
                            records.append(LinkRecord(code, full_url))
                signature = tuple(page_codes)
                if not signature:
                    raise RuntimeError(f"{category_name} 第 {page_number} 页没有 ETF 数据")
                if signature in page_signatures:
                    raise RuntimeError(f"{category_name} 翻页未更新，停止以避免遗漏")
                page_signatures.add(signature)
                emit(
                    f"上交所 {category_name} 第 {page_number} 页，"
                    f"分类累计 {len(category_codes)}/{expected_total or '?'}，"
                    f"总计 {len(records)} 条",
                    callback,
                )
                if expected_total and len(category_codes) >= expected_total:
                    break
                next_button = await page.query_selector("#tab_main li.next")
                if not next_button or not await next_button.is_visible():
                    break
                class_name = (await next_button.get_attribute("class") or "").lower()
                if "disabled" in class_name or "nonext" in class_name:
                    break
                next_link = await next_button.query_selector("a")
                if not next_link:
                    break
                previous_first_code = page_codes[0]
                await next_link.click()
                await page.wait_for_function(
                    r"""previous => {
                        const cell = document.querySelector(
                            '#tab_main table tbody tr:first-child td:nth-child(1)'
                        );
                        const match = cell && cell.textContent.match(/\d{6}/);
                        return match && match[0] !== previous;
                    }""",
                    arg=previous_first_code,
                    timeout=30_000,
                )
                await page.wait_for_timeout(600)
                page_number += 1

            if expected_total and len(category_codes) != expected_total:
                raise RuntimeError(
                    f"{category_name} 应有 {expected_total} 条，实际抓取 {len(category_codes)} 条"
                )
        await browser.close()
    return records


async def scrape_links_async(
    market: str,
    work_dir: str | os.PathLike[str] | None = None,
    *,
    headless: bool = True,
    callback: ProgressCallback | None = None,
) -> dict[str, object]:
    market = normalize_market(market)
    root = resolve_work_dir(work_dir)
    emit(f"开始抓取 {market} ETF 链接", callback)
    records = (
        await _scrape_sse(headless, callback)
        if market == "SH"
        else await _scrape_szse(headless, callback)
    )
    if not records:
        raise RuntimeError(f"{market} 未抓取到任何 ETF 链接")
    output = write_links(records, links_path(root, market))
    return {"market": market, "links": len(records), "path": str(output)}


def scrape_links(
    market: str,
    work_dir: str | os.PathLike[str] | None = None,
    *,
    headless: bool = True,
    callback: ProgressCallback | None = None,
) -> dict[str, object]:
    return asyncio.run(scrape_links_async(market, work_dir, headless=headless, callback=callback))


def _download_url(record: LinkRecord, market: str) -> str:
    if market == "SZ":
        query = parse_qs(urlparse(record.url).query)
        open_code = unquote((query.get("opencode") or [""])[0])
        filename = Path(open_code).name
        if filename:
            return urljoin(SZSE_TEXT_BASE_URL, filename)
    return record.url


def _normalize_trading_day(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 8:
        return ""
    return f"{digits[:4]}-{digits[4:6]}-{digits[6:]}"


def _record_trading_day(record: LinkRecord) -> str:
    matches = re.findall(r"(20\d{6})(?!\d)", unquote(record.url))
    return _normalize_trading_day(matches[-1]) if matches else ""


def _pcf_trading_day(path: Path, market: str) -> str:
    raw = path.read_bytes()
    text = ""
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if market == "SH":
        match = re.search(r"<TradingDay>\s*(\d{8})\s*</TradingDay>", text)
    else:
        match = re.search(r"\(\s*(\d{4}-\d{2}-\d{2})\s*\)", text)
    return _normalize_trading_day(match.group(1)) if match else ""


def _download_one(
    record: LinkRecord,
    market: str,
    destination: Path,
    force: bool,
    throttle: RequestThrottle,
) -> str:
    output = destination / f"{record.etf_code}.txt"
    if output.exists() and validate_pcf_file(output):
        if not force:
            return "skipped"
        expected_day = _record_trading_day(record)
        if expected_day and _pcf_trading_day(output, market) == expected_day:
            return "skipped"
    headers = {
        "Accept": "text/plain,text/html,application/xhtml+xml,*/*",
        "Referer": SSE_PAGE_URL if market == "SH" else SZSE_PAGE_URL,
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
        ),
    }
    session = requests.Session()
    session.trust_env = use_system_proxy()
    response = None
    for attempt in range(3):
        try:
            throttle.wait()
            response = session.get(
                _download_url(record, market),
                headers=headers,
                timeout=(15, 45),
            )
            response.raise_for_status()
            break
        except requests.RequestException as exc:
            status_code = exc.response.status_code if exc.response is not None else 0
            if status_code in {403, 429}:
                throttle.penalize(10.0 * (attempt + 1))
            if attempt >= 2:
                raise
            time.sleep(1.5 * (attempt + 1))
    assert response is not None
    content_type = response.headers.get("Content-Type", "").lower()
    sample = response.content[:200].lower()
    if "html" in content_type or b"<html" in sample:
        raise RuntimeError("下载地址返回了 HTML，可能被交易所拦截")
    if not response.content:
        raise RuntimeError("下载内容为空")
    temp_output = output.with_suffix(".txt.tmp")
    temp_output.write_bytes(response.content)
    if not validate_pcf_file(temp_output):
        temp_output.unlink(missing_ok=True)
        raise RuntimeError("下载内容不是可识别的 PCF 文件")
    temp_output.replace(output)
    return "downloaded"


def download_pcf_files(
    market: str,
    work_dir: str | os.PathLike[str] | None = None,
    *,
    force: bool = False,
    workers: int = 6,
    limit: int | None = None,
    callback: ProgressCallback | None = None,
    progress_callback: StructuredProgressCallback | None = None,
) -> dict[str, object]:
    market = normalize_market(market)
    root = resolve_work_dir(work_dir)
    records = read_links(links_path(root, market))
    if limit is not None:
        records = records[: max(0, limit)]
    destination = text_dir(root, market)
    default_delay = "0.8" if market == "SZ" else "0.15"
    delay_seconds = float(
        os.environ.get(f"PCF_{market}_REQUEST_DELAY", default_delay) or default_delay
    )
    throttle = RequestThrottle(delay_seconds)
    downloaded = 0
    skipped = 0
    errors: list[str] = []
    append_pcf_log(
        root,
        f"[{market}] 开始处理 {len(records)} 份 PCF，force={force}，"
        f"workers={workers}，request_delay={delay_seconds:.2f}s",
    )
    with ThreadPoolExecutor(max_workers=max(1, min(workers, 12))) as executor:
        futures = {
            executor.submit(
                _download_one,
                record,
                market,
                destination,
                force,
                throttle,
            ): record
            for record in records
        }
        for index, future in enumerate(as_completed(futures), start=1):
            record = futures[future]
            try:
                status = future.result()
                if status == "downloaded":
                    downloaded += 1
                else:
                    skipped += 1
            except Exception as exc:
                error = f"{record.etf_code}: {exc}"
                errors.append(error)
                append_pcf_log(root, f"[{market}] 下载失败 {error}")
            report_progress(
                progress_callback,
                phase="download",
                market=market,
                current=index,
                total=len(records),
                downloaded=downloaded,
                skipped=skipped,
                failed=len(errors),
            )
            if index == len(records) or index % 50 == 0:
                emit(
                    f"{market} PCF {index}/{len(records)}，新增 {downloaded}，失败 {len(errors)}",
                    callback,
                )
    append_pcf_log(
        root,
        f"[{market}] 处理结束：新增 {downloaded}，已有当日文件 {skipped}，"
        f"失败 {len(errors)}，总数 {len(records)}",
    )
    return {
        "market": market,
        "total": len(records),
        "downloaded": downloaded,
        "skipped": skipped,
        "failed": len(errors),
        "errors": errors[:20],
        "path": str(destination),
    }


def _decode_text(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def validate_pcf_file(path: Path) -> bool:
    try:
        if not path.exists() or path.stat().st_size <= 0:
            return False
        text = _decode_text(path)
        stripped = text.lstrip("\ufeff\r\n\t ")
        if not stripped or re.search(r"<html(?:\s|>)", stripped[:500], re.IGNORECASE):
            return False
        if stripped.startswith("<"):
            root = ElementTree.fromstring(stripped)
            component_ids = {
                "InstrumentID",
                "UnderlyingSecurityID",
            }
            return any(
                _xml_local_name(node.tag) in component_ids
                and bool((node.text or "").strip())
                for node in root.iter()
            )
        return bool(re.search(r"(?m)^\s*\d{6}(?:\s|\|)", text))
    except (OSError, ElementTree.ParseError):
        return False


def _format_sse_substitution_flag(raw_flag: str) -> str:
    flag_labels = {
        "0": "禁止",
        "1": "允许",
        "2": "必须",
        "3": "退补现金替代",
        "4": "必须现金替代",
        "5": "非沪市成分证券退补现金替代",
    }
    label = flag_labels.get(raw_flag)
    return f"{raw_flag}-{label}" if label else raw_flag


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _sse_component_market(underlying_security_id: str) -> str:
    return {
        "101": "上海市场",
        "102": "深圳市场",
        "103": "香港市场",
        "106": "北京市场",
    }.get(underlying_security_id, underlying_security_id)


def _parse_sse_components(path: Path) -> tuple[dict[str, str], list[PcfComponent]]:
    text = _decode_text(path)
    stripped = text.lstrip("\ufeff\r\n\t ")
    if not stripped.startswith("<"):
        components: list[PcfComponent] = []
        for line in text.splitlines():
            parts = [part.strip() for part in line.split("|")]
            if len(parts) < 4 or not re.fullmatch(r"\d{6}", parts[0]):
                continue
            components.append(
                PcfComponent(
                    security_code=parts[0],
                    security_name=parts[1],
                    quantity=parts[2],
                    substitute_flag=_format_sse_substitution_flag(parts[3]),
                    creation_premium_rate=parts[4] if len(parts) > 4 else "",
                    redemption_discount_rate=parts[5] if len(parts) > 5 else "",
                    substitution_cash_amount=parts[6] if len(parts) > 6 else "",
                    creation_substitution_amount="",
                    redemption_substitution_amount="",
                    component_market="",
                )
            )
        return {"trading_day": "", "record_number": str(len(components))}, components

    root = ElementTree.fromstring(stripped)
    root_fields = {
        _xml_local_name(child.tag): (child.text or "").strip()
        for child in root
        if _xml_local_name(child.tag) != "ComponentList"
    }
    components = []
    for component in root.iter():
        if _xml_local_name(component.tag) != "Component":
            continue
        fields = {
            _xml_local_name(child.tag): (child.text or "").strip()
            for child in component
        }
        security_code = fields.get("InstrumentID", "")
        if not security_code:
            continue
        raw_flag = fields.get("SubstitutionFlag", "")
        components.append(
            PcfComponent(
                security_code=security_code,
                security_name=fields.get("InstrumentName", ""),
                quantity=fields.get("Quantity", ""),
                substitute_flag=_format_sse_substitution_flag(raw_flag),
                creation_premium_rate=fields.get("CreationPremiumRate", ""),
                redemption_discount_rate=fields.get("RedemptionDiscountRate", ""),
                substitution_cash_amount=fields.get("SubstitutionCashAmount", ""),
                creation_substitution_amount="",
                redemption_substitution_amount="",
                component_market=_sse_component_market(fields.get("UnderlyingSecurityID", "")),
            )
        )
    return {
        "trading_day": root_fields.get("TradingDay", ""),
        "record_number": root_fields.get("RecordNumber", str(len(components))),
        "creation_redemption_unit": root_fields.get("CreationRedemptionUnit", ""),
    }, components


SZSE_COMPONENT_COLUMNS = (
    ("security_code", "证券代码"),
    ("security_name", "证券简称"),
    ("quantity", "股份数量"),
    ("substitute_flag", "现金替代标志"),
    ("creation_premium_rate", "申购现金替代保证金率"),
    ("redemption_discount_rate", "赎回现金替代保证金率"),
    ("creation_substitution_amount", "申购替代金额"),
    ("redemption_substitution_amount", "赎回替代金额"),
    ("component_market", "挂牌市场"),
    ("mapping_code", "映射代码"),
    ("physical_creation_redemption", "是否实物对价申赎"),
)


def _parse_szse_components(path: Path) -> tuple[dict[str, str], list[PcfComponent]]:
    raw = path.read_bytes()
    raw_lines = raw.splitlines()
    encoding = ""
    starts: list[int] = []
    for candidate in ("utf-8-sig", "gb18030"):
        labels = [label.encode(candidate) for _, label in SZSE_COMPONENT_COLUMNS]
        header = next((line for line in raw_lines if labels[0] in line), None)
        if header is None:
            continue
        candidate_starts = [header.find(label) for label in labels]
        if all(index >= 0 for index in candidate_starts):
            encoding = candidate
            starts = candidate_starts
            break
    if not encoding:
        raise ValueError(f"无法识别深市 PCF 成分列: {path.name}")

    components: list[PcfComponent] = []
    for raw_line in raw_lines:
        values: dict[str, str] = {}
        for index, (field_name, _) in enumerate(SZSE_COMPONENT_COLUMNS):
            end = starts[index + 1] if index + 1 < len(starts) else None
            values[field_name] = raw_line[starts[index]:end].decode(
                encoding, errors="ignore"
            ).strip()
        if not re.fullmatch(r"\d{6}", values["security_code"]):
            continue
        components.append(
            PcfComponent(
                security_code=values["security_code"],
                security_name=values["security_name"],
                quantity=values["quantity"],
                substitute_flag=values["substitute_flag"],
                creation_premium_rate=values["creation_premium_rate"],
                redemption_discount_rate=values["redemption_discount_rate"],
                substitution_cash_amount="",
                creation_substitution_amount=values["creation_substitution_amount"],
                redemption_substitution_amount=values["redemption_substitution_amount"],
                component_market=values["component_market"],
            )
        )
    decoded = _decode_text(path)
    trading_day_match = re.search(r"\(\s*(\d{4}-\d{2}-\d{2})\s*\)", decoded)
    return {
        "trading_day": trading_day_match.group(1) if trading_day_match else "",
        "record_number": str(len(components)),
    }, components


def get_etf_components(
    market: str,
    etf_code: str,
    work_dir: str | os.PathLike[str] | None = None,
) -> dict[str, object]:
    market = normalize_market(market)
    code = str(etf_code or "").strip()
    if not re.fullmatch(r"\d{6}", code):
        raise ValueError("ETF 代码必须是 6 位数字")
    root = resolve_work_dir(work_dir)
    source_path = text_dir(root, market) / f"{code}.txt"
    if not source_path.exists():
        raise FileNotFoundError(f"未找到 {market} ETF {code} 的 PCF 文件")
    metadata, components = (
        _parse_sse_components(source_path)
        if market == "SH"
        else _parse_szse_components(source_path)
    )
    return {
        "etf_code": code,
        "market": market,
        "source_file": source_path.name,
        "component_count": len(components),
        "metadata": metadata,
        "components": components,
    }


def _qmt_market_code(component_market: str) -> str:
    market = str(component_market or "").strip().upper()
    if market in {"SH", "SZ", "BJ", "HK"}:
        return market
    for keyword, code in (
        ("上海", "SH"),
        ("深圳", "SZ"),
        ("北京", "BJ"),
        ("香港", "HK"),
    ):
        if keyword in market:
            return code
    return market


def create_qmt_basket_csv(
    market: str,
    etf_code: str,
    work_dir: str | os.PathLike[str] | None = None,
) -> tuple[bytes, int, str]:
    result = get_etf_components(market, etf_code, work_dir)
    rows: list[list[str]] = []
    for component in result["components"]:
        try:
            quantity = Decimal(str(component.quantity).replace(",", "").strip())
        except (InvalidOperation, ValueError):
            continue
        component_market = _qmt_market_code(component.component_market)
        if quantity <= 0 or not component_market:
            continue
        quantity_text = (
            str(int(quantity))
            if quantity == quantity.to_integral_value()
            else format(quantity, "f")
        )
        rows.append(
            [component.security_code, component_market, quantity_text, "", "", "", ""]
        )

    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(["代码", "市场", "数量", "相对权重", "方向", "指定价", "量比"])
    writer.writerows(rows)
    trading_day = str(result.get("metadata", {}).get("trading_day", ""))
    return output.getvalue().encode("gb18030"), len(rows), trading_day


def _processed_components_csv(components: Sequence[PcfComponent]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow([label for _, label in PROCESSED_COMPONENT_COLUMNS])
    for component in components:
        writer.writerow(
            [getattr(component, field_name) for field_name, _ in PROCESSED_COMPONENT_COLUMNS]
        )
    return output.getvalue().encode("utf-8-sig")


def create_all_pcf_csv_zip(
    work_dir: str | os.PathLike[str] | None = None,
    *,
    markets: Sequence[str] = MARKETS,
) -> tuple[bytes, dict[str, object]]:
    root = resolve_work_dir(work_dir)
    normalized_markets = tuple(dict.fromkeys(normalize_market(market) for market in markets))
    archive_buffer = BytesIO()
    market_counts: dict[str, int] = {}
    date_counts: Counter[str] = Counter()
    file_count = 0
    component_count = 0

    with zipfile.ZipFile(
        archive_buffer,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=6,
    ) as archive:
        for market in normalized_markets:
            records = read_links(links_path(root, market))
            if not records:
                raise RuntimeError(f"{market} ETF 链接清单为空")
            market_counts[market] = 0
            for record in sorted(records, key=lambda item: item.etf_code):
                try:
                    result = get_etf_components(market, record.etf_code, root)
                except Exception as exc:
                    raise RuntimeError(
                        f"{market} ETF {record.etf_code} 的 PCF 无法导出: {exc}"
                    ) from exc
                components = result["components"]
                archive.writestr(
                    f"{market}/{record.etf_code}.csv",
                    _processed_components_csv(components),
                )
                trading_day = _normalize_trading_day(
                    str(result.get("metadata", {}).get("trading_day", ""))
                )
                if trading_day:
                    date_counts[trading_day] += 1
                market_counts[market] += 1
                file_count += 1
                component_count += len(components)

    data_date = date_counts.most_common(1)[0][0] if date_counts else ""
    return archive_buffer.getvalue(), {
        "file_count": file_count,
        "component_count": component_count,
        "market_counts": market_counts,
        "data_date": data_date,
        "date_counts": dict(sorted(date_counts.items(), reverse=True)),
    }


def _parse_sse_xml(path: Path, text: str, target_code: str) -> list[WhitelistRecord]:
    try:
        root = ElementTree.fromstring(text.lstrip("\ufeff\r\n\t "))
    except ElementTree.ParseError:
        return []

    records: list[WhitelistRecord] = []
    for component in root.iter():
        if _xml_local_name(component.tag) != "Component":
            continue
        fields = {
            _xml_local_name(child.tag): (child.text or "").strip()
            for child in component
        }
        if fields.get("InstrumentID") != target_code:
            continue
        raw_flag = fields.get("SubstitutionFlag", "")
        records.append(
            WhitelistRecord(
                path.stem,
                target_code,
                fields.get("InstrumentName", ""),
                "SH",
                _format_sse_substitution_flag(raw_flag),
                path.name,
            )
        )
    return records


def _parse_sse_text(path: Path, target_code: str) -> list[WhitelistRecord]:
    text = _decode_text(path)
    if text.lstrip("\ufeff\r\n\t ").startswith("<"):
        return _parse_sse_xml(path, text, target_code)

    records: list[WhitelistRecord] = []
    for line in text.splitlines():
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 4 or parts[0] != target_code:
            continue
        raw_flag = parts[3]
        records.append(
            WhitelistRecord(
                path.stem,
                target_code,
                parts[1] if len(parts) > 1 else "",
                "SH",
                _format_sse_substitution_flag(raw_flag),
                path.name,
            )
        )
    return records


def _parse_szse_text(path: Path, target_code: str) -> list[WhitelistRecord]:
    records: list[WhitelistRecord] = []
    pattern = re.compile(
        r"^\s*(?P<code>\d{6})\s+(?P<name>.+?)\s+[\d,]+\s+(?P<flag>允许|必须|禁止)(?:\s|$)"
    )
    for line in _decode_text(path).splitlines():
        match = pattern.match(line)
        if not match or match.group("code") != target_code:
            continue
        records.append(
            WhitelistRecord(
                path.stem,
                target_code,
                match.group("name").strip(),
                "SZ",
                match.group("flag"),
                path.name,
            )
        )
    return records


def scan_market(
    market: str,
    stock_code: str,
    work_dir: str | os.PathLike[str] | None = None,
) -> list[WhitelistRecord]:
    market = normalize_market(market)
    target_code, _ = normalize_stock_code(stock_code)
    root = resolve_work_dir(work_dir)
    parser = _parse_sse_text if market == "SH" else _parse_szse_text
    records: list[WhitelistRecord] = []
    for path in sorted(text_dir(root, market).glob("*.txt")):
        records.extend(parser(path, target_code))
    return records


def create_result_workbook(
    records: Iterable[WhitelistRecord],
    *,
    stock_name: str,
    stock_code: str,
    summaries: Sequence[dict[str, object]] = (),
    errors: Sequence[str] = (),
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ETF白名单"
    sheet.append(["ETF代码", "股票代码", "证券名称", "市场", "现金替代标志", "来源文件"])
    ordered = sorted(records, key=lambda item: (item.market, item.etf_code))
    for record in ordered:
        sheet.append(
            [
                record.etf_code,
                record.stock_code,
                record.stock_name or stock_name,
                record.market,
                record.substitute_flag,
                record.source_file,
            ]
        )
    sheet.freeze_panes = "A2"
    widths = (14, 14, 22, 10, 18, 22)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    summary_sheet = workbook.create_sheet("运行摘要")
    summary_sheet.append(["项目", "内容"])
    summary_sheet.append(["证券名称", stock_name])
    summary_sheet.append(["股票代码", stock_code])
    summary_sheet.append(["匹配ETF数量", len(ordered)])
    for summary in summaries:
        market = str(summary.get("market", ""))
        details = "，".join(f"{key}={value}" for key, value in summary.items() if key != "errors")
        summary_sheet.append([f"{market}运行结果", details])
    for error in errors:
        summary_sheet.append(["警告", error])
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 110
    return workbook


def workbook_bytes(workbook: Workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def save_market_result(
    market: str,
    stock_code: str,
    work_dir: str | os.PathLike[str] | None = None,
    output_path: str | os.PathLike[str] | None = None,
    *,
    stock_name: str = "",
) -> dict[str, object]:
    root = resolve_work_dir(work_dir)
    code, suffix = normalize_stock_code(stock_code)
    records = scan_market(market, code, root)
    destination = Path(output_path).resolve() if output_path else root / f"pcf_whitelist_{code}_{market}.xlsx"
    workbook = create_result_workbook(
        records,
        stock_name=stock_name,
        stock_code=f"{code}.{suffix}" if suffix else code,
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return {"market": market, "matches": len(records), "path": str(destination)}


def cache_status(work_dir: str | os.PathLike[str] | None = None) -> dict[str, object]:
    root = resolve_work_dir(work_dir)
    markets: dict[str, object] = {}
    for market in MARKETS:
        link_file = links_path(root, market)
        link_records: list[LinkRecord] = []
        if link_file.exists():
            try:
                link_records = read_links(link_file)
            except Exception:
                link_records = []
        records_by_code = {record.etf_code: record for record in link_records}
        link_codes = set(records_by_code)
        files = list(text_dir(root, market).glob("*.txt"))
        file_codes = {path.stem for path in files}
        valid_codes = {path.stem for path in files if validate_pcf_file(path)}
        invalid_codes = sorted((file_codes & link_codes) - valid_codes)
        missing_codes = sorted(link_codes - valid_codes)
        date_counts: Counter[str] = Counter()
        stale_codes: list[str] = []
        for code in sorted(valid_codes & link_codes):
            actual_day = _pcf_trading_day(text_dir(root, market) / f"{code}.txt", market)
            expected_day = _record_trading_day(records_by_code[code])
            if actual_day:
                date_counts[actual_day] += 1
            if expected_day and actual_day != expected_day:
                stale_codes.append(code)
        target_days = Counter(
            day for day in (_record_trading_day(record) for record in link_records) if day
        )
        if target_days:
            data_date = target_days.most_common(1)[0][0]
        elif date_counts:
            data_date = date_counts.most_common(1)[0][0]
        else:
            data_date = ""
        current_files = len(valid_codes & link_codes) - len(stale_codes)
        markets[market] = {
            "links": len(link_codes),
            "files": len(files),
            "valid_files": len(valid_codes & link_codes),
            "missing": len(missing_codes),
            "invalid": len(invalid_codes),
            "current_files": current_files,
            "stale": len(stale_codes),
            "stale_codes": stale_codes[:50],
            "data_date": data_date,
            "date_counts": dict(sorted(date_counts.items(), reverse=True)),
            "missing_codes": missing_codes[:50],
            "invalid_codes": invalid_codes[:50],
            "ready": bool(link_codes)
            and not missing_codes
            and not invalid_codes
            and not stale_codes,
        }
    return {
        "schema_version": 2,
        "work_dir": str(root),
        "log_path": str(pcf_log_path(root)),
        "markets": markets,
    }


def run_pipeline(
    stock_code: str,
    stock_name: str = "",
    work_dir: str | os.PathLike[str] | None = None,
    *,
    markets: Sequence[str] = MARKETS,
    refresh: bool = False,
    callback: ProgressCallback | None = None,
    progress_callback: StructuredProgressCallback | None = None,
) -> dict[str, object]:
    code, suffix = normalize_stock_code(stock_code)
    root = resolve_work_dir(work_dir)
    normalized_markets = tuple(dict.fromkeys(normalize_market(item) for item in markets))
    summaries: list[dict[str, object]] = []
    errors: list[str] = []
    all_records: list[WhitelistRecord] = []
    for market in normalized_markets:
        try:
            link_file = links_path(root, market)
            if refresh or not link_file.exists():
                report_progress(
                    progress_callback,
                    phase="links",
                    market=market,
                    current=0,
                    total=0,
                )
                summaries.append(scrape_links(market, root, callback=callback))
            market_status = cache_status(root)["markets"][market]
            if refresh or not market_status["ready"]:
                download_summary = download_pcf_files(
                    market,
                    root,
                    force=refresh,
                    callback=callback,
                    progress_callback=progress_callback,
                )
            else:
                total_links = len(read_links(link_file))
                download_summary = {
                    "market": market,
                    "total": total_links,
                    "downloaded": 0,
                    "skipped": int(market_status["valid_files"]),
                    "failed": 0,
                    "cache": True,
                }
                report_progress(
                    progress_callback,
                    phase="download",
                    market=market,
                    current=int(market_status["valid_files"]),
                    total=total_links,
                    downloaded=0,
                    skipped=int(market_status["valid_files"]),
                    failed=0,
                    cache=True,
                )
            summaries.append(download_summary)
            failed_downloads = int(download_summary.get("failed", 0))
            if failed_downloads:
                errors.append(
                    f"{market}: {failed_downloads} 个 PCF 文件未下载，结果可能不完整"
                )
            records = scan_market(market, code, root)
            report_progress(
                progress_callback,
                phase="scan",
                market=market,
                current=int(download_summary.get("skipped", 0))
                + int(download_summary.get("downloaded", 0)),
                total=int(download_summary.get("total", 0)),
                matches=len(records),
            )
            all_records.extend(records)
            summaries.append({"market": market, "matches": len(records)})
        except Exception as exc:
            message = f"{market}: {exc}"
            errors.append(message)
            emit(message, callback)
    if not summaries and errors:
        raise RuntimeError("；".join(errors))
    display_code = f"{code}.{suffix}" if suffix else code
    workbook = create_result_workbook(
        all_records,
        stock_name=stock_name.strip(),
        stock_code=display_code,
        summaries=summaries,
        errors=errors,
    )
    return {
        "stock_code": display_code,
        "stock_name": stock_name.strip(),
        "matches": len(all_records),
        "records": all_records,
        "summaries": summaries,
        "errors": errors,
        "workbook": workbook,
        "work_dir": str(root),
    }

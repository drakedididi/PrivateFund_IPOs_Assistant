from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence


OUTPUT_COLUMNS = [
    "配售对象全称",
    "账户估值表总资产金额（人民币：元）",
    "备注",
]

TOTAL_LABELS = ("资产类合计", "资产合计")
LOW_ASSET_THRESHOLD = Decimal("60000000")
DISPLAY_NAME_OVERRIDES = {
    "睿量兴泰锐进1号私募证券投资基金": "省心享睿量兴泰锐进1号私募证券投资基金",
}
MONEY_RE = re.compile(r"-?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?")
FUND_NAME_RE = re.compile(r"(睿量[^\\/_\s]*?基金)")


def debug_log(message: str) -> None:
    print(f"[valuation_table] {message}", flush=True)


@dataclass
class ValuationResult:
    fund_name: str
    total_assets: Decimal
    pdf_name: str


def import_pdfplumber():
    try:
        import pdfplumber  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "未安装 pdfplumber。请先运行：python -m pip install pdfplumber"
        ) from exc
    return pdfplumber


def import_openpyxl():
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "未安装 openpyxl。请先运行：python -m pip install openpyxl"
        ) from exc
    return Workbook, Alignment, Font, PatternFill, get_column_letter


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value))


def parse_money(value: object) -> Decimal | None:
    text = normalize_cell(value)
    if not text:
        return None
    match = MONEY_RE.search(text)
    if not match:
        return None
    try:
        return Decimal(match.group(0).replace(",", ""))
    except InvalidOperation:
        return None


def extract_fund_name(pdf_path: Path, root_dir: Path) -> str:
    candidates = [pdf_path.stem]
    try:
        candidates.extend(part.name for part in pdf_path.parents if part != root_dir.parent)
    except RuntimeError:
        candidates.extend(part.name for part in pdf_path.parents)

    for text in candidates:
        match = FUND_NAME_RE.search(text)
        if match:
            return match.group(1)
    raise ValueError("文件名或文件夹名中未找到“睿量...基金”格式的配售对象名称")


def row_has_total_label(row: Sequence[object]) -> bool:
    joined = "".join(normalize_cell(cell) for cell in row)
    return any(label in joined for label in TOTAL_LABELS)


def find_market_value_column(table: Sequence[Sequence[object]], row_index: int) -> int | None:
    """Find the column headed by 市值, excluding 市值占比 and 估值增值."""
    start = 0
    for header_index in range(row_index - 1, start - 1, -1):
        header_row = table[header_index]
        for index, cell in enumerate(header_row):
            text = normalize_cell(cell)
            if "市值" in text and "占" not in text and "估值" not in text:
                if has_local_currency_subcolumn(table, header_index, row_index, index):
                    return index + 1
                return index
    return None


def has_local_currency_subcolumn(
    table: Sequence[Sequence[object]],
    header_index: int,
    row_index: int,
    column_index: int,
) -> bool:
    next_column = column_index + 1
    for subheader_index in range(header_index + 1, min(row_index, header_index + 4)):
        subheader_row = table[subheader_index]
        if next_column < len(subheader_row) and normalize_cell(subheader_row[next_column]) == "本币":
            return True
    return False


def money_values_from_row(row: Sequence[object]) -> list[Decimal]:
    values: list[Decimal] = []
    for cell in row:
        amount = parse_money(cell)
        if amount is not None:
            values.append(amount)
    return values


def extract_total_assets_from_tables(pdf_path: Path) -> Decimal | None:
    pdfplumber = import_pdfplumber()

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row_index, row in enumerate(table):
                    if not row_has_total_label(row):
                        continue

                    market_value_column = find_market_value_column(table, row_index)
                    if market_value_column is not None and market_value_column < len(row):
                        amount = parse_money(row[market_value_column])
                        if amount is not None:
                            return amount

                    amounts = money_values_from_row(row)
                    if len(amounts) >= 3:
                        return amounts[2]
                    if len(amounts) == 1:
                        return amounts[0]

    return None


def extract_total_assets_from_text(pdf_path: Path) -> Decimal | None:
    pdfplumber = import_pdfplumber()

    with pdfplumber.open(str(pdf_path)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    for line in text.splitlines():
        compact = normalize_cell(line)
        if not any(label in compact for label in TOTAL_LABELS):
            continue
        amounts = [
            Decimal(match.group(0).replace(",", ""))
            for match in MONEY_RE.finditer(line)
        ]
        if len(amounts) >= 2:
            return amounts[1]
        if amounts:
            return amounts[0]
    return None


def extract_total_assets(pdf_path: Path) -> Decimal:
    amount = extract_total_assets_from_tables(pdf_path)
    if amount is None:
        amount = extract_total_assets_from_text(pdf_path)
    if amount is None:
        raise ValueError("未找到资产类合计/资产合计对应的市值字段")
    return amount


def iter_pdf_files(input_dir: Path) -> Iterable[Path]:
    return sorted(
        path for path in input_dir.rglob("*") if path.is_file() and path.suffix.lower() == ".pdf"
    )


def collect_valuation_results(input_dir: Path) -> tuple[list[ValuationResult], list[tuple[Path, str]]]:
    results: list[ValuationResult] = []
    errors: list[tuple[Path, str]] = []
    pdf_files = list(iter_pdf_files(input_dir))
    debug_log(f"pdf scan: input_dir={input_dir}, pdf_count={len(pdf_files)}")

    for index, pdf_path in enumerate(pdf_files, start=1):
        started_at = time.perf_counter()
        debug_log(f"pdf start: {index}/{len(pdf_files)} {pdf_path.name}")
        try:
            fund_name = extract_fund_name(pdf_path, input_dir)
            total_assets = extract_total_assets(pdf_path)
            results.append(
                ValuationResult(
                    fund_name=fund_name,
                    total_assets=total_assets,
                    pdf_name=pdf_path.name,
                )
            )
            debug_log(
                f"pdf ok: {index}/{len(pdf_files)} {pdf_path.name}, "
                f"fund={fund_name}, assets={total_assets}, seconds={time.perf_counter() - started_at:.2f}"
            )
        except Exception as exc:  # noqa: BLE001 - keep processing other PDFs.
            errors.append((pdf_path, str(exc)))
            debug_log(
                f"pdf failed: {index}/{len(pdf_files)} {pdf_path.name}, "
                f"seconds={time.perf_counter() - started_at:.2f}, error={exc}"
            )

    results.sort(key=lambda item: item.fund_name)
    return results, errors


def write_excel(results: Sequence[ValuationResult], output_path: Path) -> None:
    Workbook, Alignment, Font, PatternFill, get_column_letter = import_openpyxl()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "打新产品总资产"
    worksheet.append(OUTPUT_COLUMNS)

    for item in results:
        note = "总资产低于六千万" if item.total_assets < LOW_ASSET_THRESHOLD else ""
        display_name = DISPLAY_NAME_OVERRIDES.get(item.fund_name, item.fund_name)
        worksheet.append([display_name, float(item.total_assets), note])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    low_asset_fill = PatternFill("solid", fgColor="FFF2CC")
    low_asset_font = Font(color="9C5700")
    for row in worksheet.iter_rows(min_row=2, max_col=3):
        row[0].alignment = Alignment(vertical="center")
        row[1].number_format = "#,##0.00"
        row[1].alignment = Alignment(horizontal="right", vertical="center")
        row[2].alignment = Alignment(horizontal="center", vertical="center")
        if row[2].value:
            for cell in row:
                cell.fill = low_asset_fill
                cell.font = low_asset_font

    widths = [48, 40, 22]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def process_excel_files(input_dir: Path, output_dir: Path) -> dict[str, object]:
    """处理指定目录内的估值表 PDF，并把生成的 Excel 写入 output_dir。"""
    debug_log(f"processor started: input_dir={input_dir}, output_dir={output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)
    results, errors = collect_valuation_results(input_dir)
    if not results:
        details = "; ".join(f"{path.name}: {message}" for path, message in errors[:5])
        raise RuntimeError(details or "未成功解析任何 PDF")

    output_path = output_dir / "打新产品总资产.xlsx"
    write_excel(results, output_path)
    debug_log(f"excel written: path={output_path}, rows={len(results)}, errors={len(errors)}")
    return {
        "pdf_total": len(results) + len(errors),
        "pdf_processed": len(results),
        "generated_files": [str(output_path)],
        "errors": [f"{path.name}: {message}" for path, message in errors],
    }


def resolve_input_dir(value: str | None) -> Path:
    if value:
        path = Path(value)
        if not path.is_absolute():
            path = Path.cwd() / path
        return path.resolve()
    return Path(__file__).resolve().parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="遍历估值表 PDF，提取资产类合计/资产合计的市值，并生成打新产品总资产 Excel。"
    )
    parser.add_argument(
        "-i",
        "--input-dir",
        help="PDF 所在文件夹。支持相对路径；默认扫描 valuation_table.py 所在文件夹。",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="打新产品总资产.xlsx",
        help="输出 Excel 路径。默认输出到 PDF 文件夹下的 打新产品总资产.xlsx。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_dir = resolve_input_dir(args.input_dir)
    if not input_dir.exists() or not input_dir.is_dir():
        print(f"输入文件夹不存在：{input_dir}", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = input_dir / output_path

    results, errors = collect_valuation_results(input_dir)
    if not results:
        print("未成功解析任何 PDF。", file=sys.stderr)
        for pdf_path, message in errors:
            print(f"- {pdf_path.name}: {message}", file=sys.stderr)
        return 1

    write_excel(results, output_path)

    print(f"已生成：{output_path}")
    print(f"成功解析：{len(results)} 个 PDF")
    if errors:
        print(f"解析失败：{len(errors)} 个 PDF", file=sys.stderr)
        for pdf_path, message in errors:
            print(f"- {pdf_path.name}: {message}", file=sys.stderr)
    return 0 if not errors else 2


if __name__ == "__main__":
    raise SystemExit(main())
